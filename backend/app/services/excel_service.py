from pathlib import Path
from uuid import uuid4
from openpyxl import Workbook, load_workbook
from app.models import Employee, FinalizationRequest
OUTPUT_DIR = Path(__file__).resolve().parents[2] / 'outputs'; OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
def export_finalization(req: FinalizationRequest) -> str:
    filename=f"allocation_{req.daily_input.shift_date}_{uuid4().hex[:8]}.xlsx"; path=OUTPUT_DIR/filename
    wb=Workbook(); ws=wb.active; ws.title='Final Allocation'
    ws.append(['Schichtdatum',req.daily_input.shift_date]); ws.append(['Status',req.allocation_result.run_status]); ws.append([])
    ws.append(['Employee ID','Badge ID','Name','User ID','Rolle','Engine Cluster','FCLM Cluster','Finger','Primary Aisles','Primary Volume','Load %','Support Target','Support Aisles','Internal Note','Status'])
    for a in req.allocation_result.assignments: ws.append([a.employee_id,a.badge_id,a.name,a.user_id,a.role,a.engine_cluster,a.fclm_cluster,a.finger,', '.join(a.primary_aisles),a.primary_volume,a.load_percent,a.support_target,', '.join(a.support_aisles),a.internal_note,a.status])
    rp=wb.create_sheet('Role Plan'); rp.append(['Engine Cluster','Suggested','Current','Min','Max','Mandatory','Editable','Type','Available Skilled','Warnings'])
    for i in req.role_plan.items: rp.append([i.engine_cluster,i.suggested_count,i.current_count,i.minimum_count,i.maximum_count,i.mandatory,str(i.editable),i.type,i.available_skilled_employees,', '.join(i.warnings)])
    vol=wb.create_sheet('Volumes'); vol.append(['Finger','Aisle','SD Volume','ND Volume','Total'])
    for v in req.daily_input.volumes: vol.append([v.finger,v.aisle,v.sd_volume,v.nd_volume,v.total_volume])
    wb.save(path); return filename
def parse_employee_workbook(file_path: Path) -> list[Employee]:
    wb=load_workbook(file_path); sh=wb.active; headers=[str(c.value).strip().lower() if c.value else '' for c in next(sh.iter_rows(min_row=1,max_row=1))]; out=[]
    for idx,row in enumerate(sh.iter_rows(min_row=2, values_only=True),1):
        vals=dict(zip(headers,row)); eid=str(vals.get('employee_id') or vals.get('id') or f'upload-{idx}'); name=str(vals.get('name') or vals.get('employee') or '')
        if name: out.append(Employee(employee_id=eid,name=name,badge_id=str(vals.get('badge_id') or ''),user_id=str(vals.get('user_id') or ''),skills=[s.strip() for s in str(vals.get('skills') or '').split(',') if s.strip()]))
    return out
