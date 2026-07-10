from pathlib import Path
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.routes import DUMMY, RATES
from app.engine.dummy_engine import calculate_role_plan, calculate_stow_utilization, validate_role_plan
from app.main import app
from app.models import AisleVolume, AllocationResult, DailyInput, FinalizationRequest, RolePlanItem
from app.services.excel_service import OUTPUT_DIR, export_finalization


def daily_input() -> DailyInput:
    return DailyInput(
        shift_date='2026-07-10',
        employees=DUMMY[:26],
        volumes=[AisleVolume(finger='A', aisle='A1', sd_volume=300, nd_volume=200, total_volume=500), AisleVolume(finger='C', aisle='C5', sd_volume=400, nd_volume=350, total_volume=750)],
        rates=RATES,
        previous_day_found=True,
    )


def test_negative_role_count_is_rejected():
    plan = calculate_role_plan(daily_input())
    plan.items[0].current_count = -1
    result = validate_role_plan(plan)
    assert not result.is_valid
    assert any('Rollenanzahl darf nicht kleiner als 0 sein' in issue.message for issue in result.issues)


def test_zero_role_count_is_accepted_when_total_still_matches():
    plan = calculate_role_plan(daily_input())
    stow = next(i for i in plan.items if i.engine_cluster == 'Stow A')
    pick = next(i for i in plan.items if i.engine_cluster == 'Pick A')
    bc = next(i for i in plan.items if i.engine_cluster == 'Stow B/C')
    moved = stow.current_count + pick.current_count
    stow.current_count = 0
    pick.current_count = 0
    bc.current_count += moved
    plan.current_positions = sum(i.current_count for i in plan.items)
    result = validate_role_plan(plan)
    assert result.is_valid


def test_pick_a_must_equal_stow_a_and_total_reflects_coupling():
    plan = calculate_role_plan(daily_input())
    stow = next(i for i in plan.items if i.engine_cluster == 'Stow A')
    pick = next(i for i in plan.items if i.engine_cluster == 'Pick A')
    before = plan.current_positions
    stow.current_count += 2
    pick.current_count = stow.current_count
    plan.current_positions = sum(i.current_count for i in plan.items)
    assert plan.current_positions == before + 4
    result = validate_role_plan(plan)
    assert any('Position' in issue.message for issue in result.issues)


def test_pick_a_independent_change_is_rejected():
    plan = calculate_role_plan(daily_input())
    next(i for i in plan.items if i.engine_cluster == 'Pick A').current_count += 1
    result = validate_role_plan(plan)
    assert not result.is_valid
    assert any('Pick A muss automatisch Stow A entsprechen' in issue.message for issue in result.issues)


def test_additional_role_can_be_added_and_counts_in_validation():
    plan = calculate_role_plan(daily_input())
    plan.items.append(RolePlanItem(engine_cluster='Problem Solver', suggested_count=0, current_count=1, type='optional', available_skilled_employees=999))
    plan.current_positions = sum(i.current_count for i in plan.items)
    result = validate_role_plan(plan)
    assert not result.is_valid
    assert result.current_positions == plan.target_positions + 1


def test_duplicate_role_is_rejected():
    plan = calculate_role_plan(daily_input())
    plan.items.append(plan.items[0].model_copy(deep=True))
    result = validate_role_plan(plan)
    assert not result.is_valid
    assert any('doppelt' in issue.message for issue in result.issues)


def test_stow_utilization_changes_with_stower_count():
    inp = daily_input()
    util_8 = calculate_stow_utilization('Stow B/C', 8, inp)
    util_7 = calculate_stow_utilization('Stow B/C', 7, inp)
    assert util_7 and util_8
    assert util_7 > util_8


def test_finalize_creates_excel_file_with_required_sheets():
    inp = daily_input()
    plan = calculate_role_plan(inp)
    result = AllocationResult(run_status='dummy_success', kpis={}, assignments=[])
    filename = export_finalization(FinalizationRequest(daily_input=inp, role_plan=plan, allocation_result=result))
    path = OUTPUT_DIR / filename
    assert path.exists()
    wb = load_workbook(path)
    assert {'Finale Allocation', 'Rollenplan', 'Volumendaten', 'Inputs'}.issubset(set(wb.sheetnames))


def test_export_endpoint_serves_xlsx_and_missing_file_is_clear():
    client = TestClient(app)
    inp = daily_input()
    plan = calculate_role_plan(inp)
    filename = export_finalization(FinalizationRequest(daily_input=inp, role_plan=plan, allocation_result=AllocationResult(run_status='dummy_success', kpis={}, assignments=[])))
    response = client.get(f'/api/exports/{filename}')
    assert response.status_code == 200
    assert response.headers['content-type'].startswith('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    missing = client.get('/api/exports/does-not-exist.xlsx')
    assert missing.status_code == 404
    assert 'wurde nicht gefunden' in missing.text
